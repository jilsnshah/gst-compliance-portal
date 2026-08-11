from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Optional

from openpyxl import load_workbook

from app.core.enums import InvoiceSource

# Canonical field -> header aliases. Client purchase registers and GST portal
# 2B exports never agree on wording, so mapping is config, not code.
COLUMN_ALIASES = {
    "supplier_gstin": [
        "gstin", "gstin of supplier", "supplier gstin", "gstin/uin of recipient",
        "supplier gstin/uin", "gstin of the supplier", "party gstin", "vendor gstin",
    ],
    "supplier_name": [
        "trade/legal name", "supplier name", "trade name", "legal name",
        "party name", "vendor name", "name of supplier",
    ],
    "invoice_no": [
        "invoice number", "invoice no", "invoice no.", "inv no", "bill no",
        "document number", "doc no", "invoice/document no", "bill number",
    ],
    "invoice_date": [
        "invoice date", "inv date", "bill date", "document date", "doc date", "date",
    ],
    "invoice_type": ["invoice type", "document type", "type", "nature of supply"],
    "place_of_supply": ["place of supply", "pos", "state"],
    "taxable_value": [
        "taxable value", "taxable value (rs)", "taxable amount", "assessable value",
        "basic amount", "net amount", "taxable val",
    ],
    "igst": ["integrated tax", "igst", "igst amount", "integrated tax(rs)", "igst (rs)"],
    "cgst": ["central tax", "cgst", "cgst amount", "central tax(rs)", "cgst (rs)"],
    "sgst": [
        "state/ut tax", "state tax", "sgst", "sgst amount", "sgst/utgst",
        "state/ut tax(rs)", "sgst (rs)",
    ],
    "cess": ["cess", "cess amount", "cess(rs)"],
    "total_value": [
        "invoice value", "total value", "total amount", "invoice value (rs)",
        "gross total", "bill amount",
    ],
    "itc_available": ["itc availability", "itc available", "availability of itc"],
}

NUMERIC_FIELDS = ["taxable_value", "igst", "cgst", "sgst", "cess", "total_value"]
REQUIRED_FIELDS = ["supplier_gstin", "invoice_no", "invoice_date", "taxable_value"]

_HEADER_LOOKUP = {}
for field, aliases in COLUMN_ALIASES.items():
    for alias in aliases:
        _HEADER_LOOKUP[alias] = field


def _norm_header(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[\n\r]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" :*")


def normalize_invoice_no(value) -> str:
    """Join key: upper-cased, punctuation removed, leading zeros dropped.
    'INV-0012/26' and 'inv 12/26' collapse to the same key."""
    if value is None:
        return ""
    text = re.sub(r"[^A-Za-z0-9]", "", str(value)).upper()
    stripped = text.lstrip("0")
    return stripped or text


def parse_date(value) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d %b %Y", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_number(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = re.sub(r"[,\s₹]", "", str(value))
    if text in ("", "-"):
        return 0.0
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    try:
        num = float(text)
    except ValueError:
        return 0.0
    return round(-num if negative else num, 2)


def _find_header_row(rows, max_scan: int = 15):
    """GST portal exports carry preamble rows, so the header is located by
    scoring each candidate row against the alias table."""
    best_index, best_map, best_score = None, {}, 0
    for idx, row in enumerate(rows[:max_scan]):
        mapping, score = {}, 0
        for col_idx, cell in enumerate(row):
            field = _HEADER_LOOKUP.get(_norm_header(cell))
            if field and field not in mapping:
                mapping[field] = col_idx
                score += 1
        if score > best_score:
            best_index, best_map, best_score = idx, mapping, score
    return best_index, best_map, best_score


# Fields a mapping may address. Only the required ones must be present.
MAPPABLE_FIELDS = [
    "supplier_gstin",
    "supplier_name",
    "invoice_no",
    "invoice_date",
    "invoice_type",
    "place_of_supply",
    "taxable_value",
    "igst",
    "cgst",
    "sgst",
    "cess",
    "total_value",
    "itc_available",
]


def column_letter(index: int) -> str:
    """0 -> A, 25 -> Z, 26 -> AA. What the CA sees in Excel."""
    letters = ""
    index += 1
    while index:
        index, rem = divmod(index - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def header_fingerprint(row) -> str:
    """Stable hash of a header row, so a saved mapping can be recognised next
    month and a changed export format can be spotted instead of mis-parsed."""
    import hashlib

    joined = "|".join(_norm_header(c) for c in row)
    return hashlib.sha256(joined.encode()).hexdigest()[:32]


class UnreadableFile(Exception):
    """The upload is not a workbook we can read. Carries a message aimed at
    whoever has to fix it, not a zipfile stack trace."""


def sniff_format(data: bytes, filename: str = "") -> str:
    name = (filename or "").lower()
    if data[:5] == b"%PDF-":
        return "pdf"
    if data[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "xls"
    if data[:4] == b"PK\x03\x04":
        # Every modern office format is a zip. Look inside to tell them apart.
        try:
            import zipfile

            names = zipfile.ZipFile(io.BytesIO(data)).namelist()
        except Exception:
            return "zip"
        if any(n.startswith("xl/") for n in names):
            return "xlsx"
        if any(n.startswith("Index/") or n.endswith(".iwa") for n in names):
            return "numbers"
        if "mimetype" in names:
            return "ods"
        return "zip"
    if name.endswith(".csv") or _looks_like_csv(data):
        return "csv"
    return "unknown"


def _looks_like_csv(data: bytes) -> bool:
    head = data[:2048]
    if b"\x00" in head:
        return False
    try:
        text = head.decode("utf-8-sig")
    except UnicodeDecodeError:
        return False
    lines = [ln for ln in text.splitlines() if ln.strip()][:5]
    return len(lines) >= 2 and all(("," in ln or "\t" in ln) for ln in lines)


UNREADABLE = {
    "numbers": "This is an Apple Numbers file. In Numbers choose File \u2192 Export To "
               "\u2192 Excel, and upload the .xlsx.",
    "xls": "This is an old .xls workbook. Open it and use Save As \u2192 .xlsx, "
           "or export it as CSV.",
    "ods": "This is an OpenDocument spreadsheet. Save it as .xlsx or CSV.",
    "pdf": "This is a PDF, not a spreadsheet. Invoice data has to come as a "
           "workbook or CSV so it can be read line by line.",
    "zip": "This is a zip archive, not a spreadsheet.",
    "unknown": "This file is not a spreadsheet we can read. Upload .xlsx or CSV.",
}


def _csv_rows(data: bytes) -> list:
    import csv as _csv

    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = _csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except Exception:
        dialect = _csv.excel
    return [list(r) for r in _csv.reader(io.StringIO(text), dialect)]


def _load_rows(data: bytes, filename: str = "", sheet_name: Optional[str] = None):
    """Returns (sheet_names, rows). One code path for xlsx and CSV so the
    mapping UI behaves the same either way."""
    kind = sniff_format(data, filename)
    if kind == "csv":
        return ["CSV"], _csv_rows(data)
    if kind != "xlsx":
        raise UnreadableFile(UNREADABLE.get(kind, UNREADABLE["unknown"]))
    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    names = workbook.sheetnames
    target = sheet_name if sheet_name in names else names[0]
    rows = [list(r) for r in workbook[target].iter_rows(values_only=True)]
    workbook.close()
    return names, rows


def preview_workbook(data: bytes, max_rows: int = 15) -> dict:
    """Every sheet's first rows, verbatim, addressed by column letter.

    Deliberately makes no attempt to find a header or identify fields: the
    point is to show the CA exactly what is in the file so they can say which
    column is which."""
    names, _ = _load_rows(data)
    sheets = []
    for name in names:
        _, raw = _load_rows(data, sheet_name=name)
        rows = []
        for i, row in enumerate(raw):
            if i >= max_rows:
                break
            rows.append(
                [
                    "" if c is None else (c.isoformat() if hasattr(c, "isoformat") else str(c))
                    for c in row
                ]
            )
        width = max((len(r) for r in rows), default=0)
        rows = [r + [""] * (width - len(r)) for r in rows]
        sheets.append(
            {
                "name": name,
                "columns": [column_letter(i) for i in range(width)],
                "rows": rows,
                "suggested_header_row": _guess_header_row(rows),
            }
        )
    return {"sheets": sheets, "fields": MAPPABLE_FIELDS, "required": REQUIRED_FIELDS}


def _guess_header_row(rows) -> Optional[int]:
    """A hint only -- the CA confirms or overrides it. Returns a 1-based row."""
    best, best_score = None, 0
    for idx, row in enumerate(rows[:15]):
        score = sum(1 for c in row if _HEADER_LOOKUP.get(_norm_header(c)))
        if score > best_score:
            best, best_score = idx + 1, score
    return best if best_score >= 3 else None


def parse_with_mapping(data: bytes, mapping: dict) -> dict:
    """Parses using column positions the CA stated. No guessing anywhere.

    mapping = {sheet_name, header_row (1-based or None), first_data_row
    (1-based), columns: {field: 0-based index}}
    """
    try:
        names, rows = _load_rows(data, sheet_name=mapping.get("sheet_name"))
    except UnreadableFile as exc:
        return {
            "records": [], "errors": [str(exc)], "warnings": [],
            "header_row": None, "mapped_columns": {}, "unmapped_headers": [],
        }
    name = mapping.get("sheet_name") or names[0]
    if name not in names:
        return {
            "records": [],
            "errors": [f"Sheet '{name}' is not in this file"],
            "warnings": [],
            "header_row": None,
            "mapped_columns": {},
            "unmapped_headers": [],
        }

    columns = {f: int(i) for f, i in (mapping.get("columns") or {}).items() if i is not None}
    missing = [f for f in REQUIRED_FIELDS if f not in columns]
    errors = []
    if missing:
        errors.append("Mapping is missing required columns: " + ", ".join(missing))
        return {
            "records": [],
            "errors": errors,
            "header_row": mapping.get("header_row"),
            "mapped_columns": columns,
            "unmapped_headers": [],
        }

    start = max(int(mapping.get("first_data_row") or 1) - 1, 0)
    records = []
    for offset, row in enumerate(rows[start:], start=start + 1):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        rec, row_errors = _row_to_record(row, columns, offset)
        if row_errors:
            errors.append(f"Row {offset}: " + "; ".join(row_errors))
            if not rec["invoice_no"]:
                continue
        records.append(rec)

    return {
        "records": records,
        "errors": errors,
        "warnings": validate_mapping(records, columns),
        "header_row": mapping.get("header_row"),
        "mapped_columns": columns,
        "unmapped_headers": [],
    }


GSTIN_RE = re.compile(r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][0-9A-Z]{3}$")


def validate_mapping(records, columns) -> list:
    """Does the mapped data actually look like what it claims to be?

    A mapping that points at the wrong columns still parses -- GSTINs land in
    the invoice-number field and text amounts quietly become 0.00 -- so the
    shape of the values is checked before any of it reaches reconciliation."""
    if not records:
        return []
    warnings = []
    total = len(records)

    bad_gstin = sum(1 for r in records if not GSTIN_RE.match(r.get("supplier_gstin") or ""))
    if bad_gstin > total / 2:
        sample = next((r["supplier_gstin"] for r in records if r.get("supplier_gstin")), "blank")
        warnings.append(
            f"{bad_gstin} of {total} values in the supplier GSTIN column are not GSTINs "
            f"(e.g. {sample!r}) -- is that the right column?"
        )

    zero_value = sum(1 for r in records if not r.get("taxable_value"))
    if zero_value > total / 2:
        warnings.append(
            f"{zero_value} of {total} rows have a taxable value of 0 -- the taxable value "
            "column may be pointing at text rather than numbers"
        )

    no_date = sum(1 for r in records if r.get("invoice_date") is None)
    if no_date > total / 2:
        warnings.append(f"{no_date} of {total} rows have no readable invoice date")

    if columns.get("invoice_no") is not None:
        looks_gstin = sum(1 for r in records if GSTIN_RE.match((r.get("invoice_no") or "").upper()))
        if looks_gstin > total / 2:
            warnings.append(
                "the invoice number column looks like it contains GSTINs -- "
                "the columns may be swapped"
            )
    return warnings


def _row_to_record(row, columns: dict, row_no: int):
    def cell(field):
        idx = columns.get(field)
        return row[idx] if idx is not None and idx < len(row) else None

    rec = {
        "source_row_no": row_no,
        "supplier_gstin": (str(cell("supplier_gstin")).strip().upper()
                           if cell("supplier_gstin") else None),
        "supplier_name": str(cell("supplier_name")).strip() if cell("supplier_name") else None,
        "invoice_no": str(cell("invoice_no")).strip() if cell("invoice_no") else None,
        "invoice_date": parse_date(cell("invoice_date")),
        "invoice_type": str(cell("invoice_type")).strip() if cell("invoice_type") else None,
        "place_of_supply": (str(cell("place_of_supply")).strip()
                            if cell("place_of_supply") else None),
        "itc_available": (str(cell("itc_available")).strip()[:5]
                          if cell("itc_available") else None),
        "raw": {},
    }
    for field in NUMERIC_FIELDS:
        rec[field] = parse_number(cell(field))
    rec["invoice_no_normalized"] = normalize_invoice_no(rec["invoice_no"])
    if not rec["total_value"]:
        rec["total_value"] = round(
            rec["taxable_value"] + rec["igst"] + rec["cgst"] + rec["sgst"] + rec["cess"], 2
        )

    errors = []
    if not rec["invoice_no"]:
        errors.append("no invoice number in the mapped column")
    if not rec["supplier_gstin"]:
        errors.append("no supplier GSTIN in the mapped column")
    if rec["invoice_date"] is None:
        errors.append("unreadable invoice date")
    return rec, errors


def parse_invoice_workbook(data: bytes, source: InvoiceSource) -> dict:
    """Returns {'records': [...], 'errors': [...], 'header_row': int,
    'mapped_columns': {...}, 'unmapped_headers': [...]}"""
    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    workbook.close()

    if not rows:
        return {"records": [], "errors": ["Workbook is empty"], "header_row": None,
                "mapped_columns": {}, "unmapped_headers": []}

    header_index, mapping, score = _find_header_row(rows)
    if header_index is None or score < 3:
        return {
            "records": [],
            "errors": [
                "Could not identify a header row. Expected columns such as "
                "GSTIN, Invoice Number, Invoice Date, Taxable Value."
            ],
            "header_row": None,
            "mapped_columns": {},
            "unmapped_headers": [],
        }

    header_row = rows[header_index]
    unmapped = [
        str(c) for i, c in enumerate(header_row)
        if c and i not in mapping.values() and _norm_header(c)
    ]

    records, errors = [], []
    missing = [f for f in REQUIRED_FIELDS if f not in mapping]
    if missing:
        errors.append("Missing required columns: " + ", ".join(missing))

    for offset, row in enumerate(rows[header_index + 1:], start=header_index + 2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        def cell(field):
            idx = mapping.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        rec = {
            "source_row_no": offset,
            "supplier_gstin": (str(cell("supplier_gstin")).strip().upper()
                               if cell("supplier_gstin") else None),
            "supplier_name": str(cell("supplier_name")).strip() if cell("supplier_name") else None,
            "invoice_no": str(cell("invoice_no")).strip() if cell("invoice_no") else None,
            "invoice_date": parse_date(cell("invoice_date")),
            "invoice_type": str(cell("invoice_type")).strip() if cell("invoice_type") else None,
            "place_of_supply": (str(cell("place_of_supply")).strip()
                                if cell("place_of_supply") else None),
            "itc_available": (str(cell("itc_available")).strip()[:5]
                              if cell("itc_available") else None),
        }
        for field in NUMERIC_FIELDS:
            rec[field] = parse_number(cell(field))

        rec["invoice_no_normalized"] = normalize_invoice_no(rec["invoice_no"])

        if not rec["total_value"]:
            rec["total_value"] = round(
                rec["taxable_value"] + rec["igst"] + rec["cgst"] + rec["sgst"] + rec["cess"], 2
            )

        row_errors = []
        if not rec["invoice_no"]:
            row_errors.append("missing invoice number")
        if not rec["supplier_gstin"]:
            row_errors.append("missing supplier GSTIN")
        if rec["invoice_date"] is None:
            row_errors.append("unreadable invoice date")
        if row_errors:
            errors.append(f"Row {offset}: " + "; ".join(row_errors))
            if not rec["invoice_no"]:
                continue

        rec["raw"] = {
            _norm_header(header_row[i]): (str(row[i]) if row[i] is not None else None)
            for i in range(min(len(header_row), len(row)))
            if header_row[i] is not None
        }
        records.append(rec)

    return {
        "records": records,
        "errors": errors,
        "header_row": header_index + 1,
        "mapped_columns": {k: v for k, v in mapping.items()},
        "unmapped_headers": unmapped,
        "source": source.value,
    }
